from openpyxl import load_workbook
# set file path
filepath="/Hot_Spot_Data.xlsx"
# load file
wb=load_workbook(filepath)
# select demo.xlsx
# sheet=wb.active
sheet=wb['Raw']
# get max row count
max_row = sheet.max_row
# get max column count
max_column = sheet.max_column
if not 'Out 1' in wb.sheetnames:
    print('Out 1 doesnt exist')
    wb.create_sheet('Out 1')
else:
    wb.remove(wb['Out 1'])
    wb.create_sheet('Out 1')
sheet2 = wb['Out 1']

# accomadate for multiple offers, kick out all strikethroughs, dump all into rowlist
####################################################
# iterate over all cells
# iterate over all rows
rowlist = []
for i in range(1, max_row + 1):

    contentslist = []
    # iterate over all columns
    for j in range(1, max_column + 1):
        # get particular cell value
        cell_obj = sheet.cell(row=i, column=j)
        cell_out = sheet2.cell(row=i, column=j)
        # print cell value
        if not cell_obj.font.strike:
            cell_obj = str(cell_obj.value).replace(" ","").replace("\n","").replace(",","")
            # print(cell_obj, end=' | ')
            # print(len(str(cell_obj)), end = "")
            cell_out.value = cell_obj.replace("/","")
            if not j == max_column:
                if cell_obj == "Pg3":
                    cell_obj = "3"
                contentslist.append(cell_obj)
            elif not contentslist[0].find("/") == -1 and j == max_column: # if there is a "/" in the first column, and the iteration is at column 5
                cell_obj = cell_obj.replace("/","")
                contentslist.append(cell_obj)
                contentslist2 = [] + contentslist
                contentslist[0] = contentslist[0][:3]
                contentslist2[0] = contentslist2[0][4:]
                if len(contentslist) == 5:
                    rowlist.append(contentslist)
                if len(contentslist2) == 5:
                    rowlist.append(contentslist2)
                contentslist = []
            elif j == max_column:
                cell_obj = cell_obj.replace("/", "")
                contentslist.append(cell_obj)
                if len(contentslist) == 5:
                    rowlist.append(contentslist)
                contentslist = []

# Dump contents of rowlist into ['Out 1']
##########################################################################

for i in range(0, len(rowlist)):



    #interate over all columns
    for j in range(0, len(rowlist[0])):
        if len(rowlist[i]) == 5:
            cell_out = sheet2.cell(row=i+1, column=j+1)
            cell_out.value = rowlist[i][j]

########################################################################################################################
                                            #Process Out1 -> Dump to Out 2
########################################################################################################################
#1. Process


if not 'Out 2' in wb.sheetnames:
    print('Out 2 doesnt exist')
    wb.create_sheet('Out 2')
else:
    wb.remove(wb['Out 2'])
    wb.create_sheet('Out 2')
sheet3 = wb['Out 2']
# for row in sheet3['A1:Z999999']:
#   for cell in row:
#     cell.value = None
# get max row count
max_row = sheet2.max_row
# get max column count
max_column = sheet2.max_column

# iterate over all cells
# iterate over all rows
rowlist = []
for i in range(1, max_row + 1):
    #
    contentslist = []
    # iterate over all columns
    for j in range(1, max_column + 1):
        # get particular cell value
        cell_out1 = sheet2.cell(row=i, column=j)
        # cell_out2 = sheet3.cell(row=i, column=j)


        cell_out1 = cell_out1.value


        contentslist.append(cell_out1)

        if j == max_column and len(contentslist[max_column-1]) > 5:
            if len(contentslist[max_column-1]) % 5 == 0:
                if len(contentslist[max_column-1]) >= 10:
                    contentslist2 = [] + contentslist
                    contentslist1 = [] + contentslist
                    contentslist1[4] = contentslist[4][:5]
                    contentslist2[4] = contentslist[4][5:10]
                    rowlist.append(contentslist1)
                    rowlist.append(contentslist2)
                    if len(contentslist[max_column - 1]) >= 15:
                        contentslist3 = [] + contentslist
                        contentslist3[4] = contentslist[4][10:15]
                        rowlist.append(contentslist3)
                        if len(contentslist[max_column - 1]) >= 20:
                            contentslist4 = [] + contentslist
                            contentslist4[4] = contentslist[4][15:20]
                            rowlist.append(contentslist4)
                            if len(contentslist[max_column - 1]) >= 25:
                                contentslist5 = [] + contentslist
                                contentslist5[4] = contentslist[4][20:25]
                                rowlist.append(contentslist5)
                                if len(contentslist[max_column - 1]) >= 30:
                                    contentslist6 = [] + contentslist
                                    contentslist6[4] = contentslist[4][25:30]
                                    rowlist.append(contentslist6)
                                    if len(contentslist[max_column - 1]) >= 35:
                                        contentslist7 = [] + contentslist
                                        contentslist7[4] = contentslist[4][30:35]
                                        rowlist.append(contentslist7)
                                        if len(contentslist[max_column - 1]) >= 40:
                                            contentslist8 = [] + contentslist
                                            contentslist8[4] = contentslist[4][35:40]
                                            rowlist.append(contentslist8)
                                            if len(contentslist[max_column - 1]) >= 45:
                                                contentslist9 = [] + contentslist
                                                contentslist9[4] = contentslist[4][40:45]
                                                rowlist.append(contentslist9)
                                                if len(contentslist[max_column - 1]) >= 50:
                                                    contentslist10 = [] + contentslist
                                                    contentslist10[4] = contentslist[4][45:50]
                                                    rowlist.append(contentslist10)
                                                    if len(contentslist[max_column - 1]) >= 55:
                                                        contentslist11 = [] + contentslist
                                                        contentslist11[4] = contentslist[4][50:55]
                                                        rowlist.append(contentslist11)
                                                        if len(contentslist[max_column - 1]) >= 60:
                                                            contentslist12 = [] + contentslist
                                                            contentslist12[4] = contentslist[4][55:60]
                                                            rowlist.append(contentslist12)
                                                            if len(contentslist[max_column - 1]) >= 65:
                                                                contentslist13 = [] + contentslist
                                                                contentslist13[4] = contentslist[4][60:65]
                                                                rowlist.append(contentslist13)
                                                                if len(contentslist[max_column - 1]) >= 70:
                                                                    contentslist14 = [] + contentslist
                                                                    contentslist14[4] = contentslist[4][65:70]
                                                                    rowlist.append(contentslist14)
                                                                    if len(contentslist[max_column - 1]) >= 75:
                                                                        contentslist15 = [] + contentslist
                                                                        contentslist15[4] = contentslist[4][70:75]
                                                                        rowlist.append(contentslist15)
                                                                        if len(contentslist[max_column - 1]) >= 80:
                                                                            contentslist16 = [] + contentslist
                                                                            contentslist16[4] = contentslist[4][75:80]
                                                                            rowlist.append(contentslist16)
                                                                            if len(contentslist[max_column - 1]) >= 85:
                                                                                contentslist17 = [] + contentslist
                                                                                contentslist17[4] = contentslist[4][80:85]
                                                                                rowlist.append(contentslist17)
                                                                                if len(contentslist[max_column - 1]) >= 90:
                                                                                    contentslist18 = [] + contentslist
                                                                                    contentslist18[4] = contentslist[4][85:90]
                                                                                    rowlist.append(contentslist18)
                                                                                    if len(contentslist[max_column - 1]) >= 95:
                                                                                        contentslist19 = [] + contentslist
                                                                                        contentslist19[4] = contentslist[4][90:95]
                                                                                        rowlist.append(contentslist19)
                                                                                        if len(contentslist[max_column - 1]) >= 100:
                                                                                            contentslist20 = [] + contentslist
                                                                                            contentslist20[4] = contentslist[4][95:100]
                                                                                            rowlist.append(contentslist20)
            elif contentslist[max_column-1] == "OfferProductId":
                rowlist.append(contentslist)
            contentslist = []
            contentslist1 = []
            contentslist2 = []
            contentslist3 = []
            contentslist4 = []
            contentslist5 = []
            contentslist6 = []
            contentslist7 = []
            contentslist8 = []
            contentslist9 = []
            contentslist10 = []
            contentslist11 = []
            contentslist12 = []
            contentslist13 = []
            contentslist14 = []
            contentslist15 = []
            contentslist16 = []
            contentslist17 = []
            contentslist18 = []
            contentslist19 = []
            contentslist20 = []
        elif j == max_column:
            rowlist.append(contentslist)
            contentslist = []


########################################################################################################################
#2. Dump


for i in range(0, len(rowlist)):



    #interate over all columns

    for j in range(0, len(rowlist[0])):
        if len(rowlist[i]) == 5:
            cell_out2 = sheet3.cell(row=i+1, column=j+1)
            cell_out2.value = rowlist[i][j]
wb.save(filepath)
